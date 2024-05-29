import flet as ft
from openpyxl import load_workbook, Workbook
from flet_core.types import MainAxisAlignment
def main(page: ft.Page):
    page.window_width = 400
    page.window_height = 700
    user_inputs_container = ft.ListView(expand=1, auto_scroll=False, padding=10)
    def button_clicked(e):
        t.value = f"Entry will be of {dd.value} block"
        page.update()
    t = ft.Text()
    b = ft.ElevatedButton(text="Submit", on_click=button_clicked)

    dd = ft.Dropdown(
        width=150,
        options=[
            ft.dropdown.Option("B.Tech"),
            ft.dropdown.Option("BCA"),
            ft.dropdown.Option("BBA"),
            ft.dropdown.Option("M.Tech"),
            ft.dropdown.Option("MCA"),
            ft.dropdown.Option("MBA")
        ],
    )
    rowItem = ft.Row(controls=[dd, b], alignment=MainAxisAlignment.SPACE_EVENLY)
    page.add(rowItem, t)
    labels = ["Sl No.", "Name", "Enrollment Number", "Course", "Semester", "Year", "Specialization", "Batch","Mobile Number","Address"]
    user_inputs = []
    for i in range(len(labels)):
        user_inputs.append(ft.TextField(label=labels[i], width=100))
        user_inputs_container.controls.append(user_inputs[i])
        if i < 9:
            user_inputs_container.controls.append(ft.Container(height=30))
    def export_data(e):
        data = [input_field.value for input_field in user_inputs]
        try:
            wb = load_workbook(filename="D:/Flet_App/Student_Details.xlsx")
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(labels)
        ws.append(data)
        wb.save("E:\My_Projects\Flet_App\Std_Details.xlsx")
        page.show_snack_bar(ft.Text("Data exported to Std_Details.xlsx"))
    export_button = ft.ElevatedButton(text="Export", on_click=export_data)
    page.add(user_inputs_container, export_button)
ft.app(target=main)